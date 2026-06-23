import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def categorize_transaction(desc):
    """
    Categorizes a transaction based on keywords in its description.
    """
    desc = str(desc).lower()
    
    categories = {
        'Income/Salary': ['payroll', 'salary', 'deposit', 'direct dep', 'transfer in', 'wire in', 'refund', 'interest payment', 'dividends'],
        'Groceries': ['grocery', 'safeway', 'kroger', 'whole foods', 'trader joe', 'supermarket', 'walmart', 'target', 'costco', 'heeb', 'aldi'],
        'Restaurants/Dining': ['restaurant', 'cafe', 'starbucks', 'coffee', 'mcdonald', 'grubhub', 'ubereats', 'pizza', 'diner', 'bar', 'pub', 'grill', 'bistro'],
        'Utilities': ['electric', 'water', 'gas', 'power', 'comcast', 'verizon', 'att', 'internet', 'utility', 'trash', 'sewer', 'heating', 'netflix', 'spotify', 'hulu', 'subscriptions'],
        'Housing/Rent': ['rent', 'mortgage', 'housing', 'landlord', 'hoa', 'lease'],
        'Transfers/Investing': ['transfer', 'venmo', 'paypal', 'zelle', 'schwab', 'fidelity', 'vanguard', 'robinhood', 'atm withdrawal', 'cash withdrawal'],
        'Transportation/Auto': ['gasoline', 'chevron', 'shell', 'exxon', 'mobil', 'uber', 'lyft', 'subway', 'transit', 'auto', 'car wash', 'parking', 'toll'],
        'Shopping/Retail': ['amazon', 'ebay', 'macys', 'nordstrom', 'clothing', 'retail', 'electronics', 'apple', 'best buy', 'nike'],
        'Insurance/Medical': ['insurance', 'geico', 'progressive', 'medical', 'doctor', 'copay', 'pharmacy', 'cvs', 'walgreens', 'dentist', 'hospital'],
    }
    
    for category, keywords in categories.items():
        if any(kw in desc for kw in keywords):
            return category
            
    return 'Miscellaneous'

def export_to_excel(df_reconciled, summary, output_path):
    """
    Exports reconciled transaction dataframe and summary metadata into a styled Excel workbook.
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Dashboard / Monthly Summary
    # ----------------------------------------------------
    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Color palette
    navy_dark = "1F4E78"
    navy_light = "DDEBF7"
    white = "FFFFFF"
    grey_light = "F2F2F2"
    green_fill = "C6EFCE"
    green_font = "006100"
    red_fill = "FFC7CE"
    red_font = "9C0006"
    
    # Title
    ws_dash['A1'] = "FINANCIAL STATEMENT ANALYSIS"
    ws_dash['A1'].font = Font(name="Calibri", size=18, bold=True, color="1F4E78")
    
    ws_dash['A2'] = f"Account Type: {summary['account_type']}"
    ws_dash['A2'].font = Font(name="Calibri", size=11, italic=True)
    
    # Draw KPI Blocks
    kpis = [
        ("Starting Balance", summary['starting_balance']),
        ("Total Deposits (Credits)", summary['total_credits']),
        ("Total Withdrawals (Debits)", summary['total_debits']),
        ("Net Cash Flow", summary['net_cash_flow']),
        ("Ending Balance", summary['ending_balance'])
    ]
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for idx, (label, val) in enumerate(kpis):
        col_idx = idx + 1
        col_letter = get_column_letter(col_idx)
        
        # Label cell (Row 4)
        lbl_cell = ws_dash[f"{col_letter}4"]
        lbl_cell.value = label
        lbl_cell.font = Font(name="Calibri", size=9, bold=True, color="595959")
        lbl_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lbl_cell.border = thin_border
        
        # Value cell (Row 5)
        val_cell = ws_dash[f"{col_letter}5"]
        val_cell.value = val
        val_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        val_cell.number_format = "$#,##0.00;($#,##0.00);\"-\""
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.border = thin_border
        
    # Reconciled status badge
    ws_dash['F4'] = "Reconciliation Status"
    ws_dash['F4'].font = Font(name="Calibri", size=9, bold=True, color="595959")
    ws_dash['F4'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws_dash['F4'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_dash['F4'].border = thin_border
    
    status_cell = ws_dash['F5']
    if summary['is_reconciled']:
        status_cell.value = "RECONCILED"
        status_cell.font = Font(name="Calibri", size=12, bold=True, color=green_font)
        status_cell.fill = PatternFill(start_color=green_fill, end_color=green_fill, fill_type="solid")
    else:
        status_cell.value = f"MISMATCH ({summary['discrepancies_count']} Err)"
        status_cell.font = Font(name="Calibri", size=11, bold=True, color=red_font)
        status_cell.fill = PatternFill(start_color=red_fill, end_color=red_fill, fill_type="solid")
    status_cell.alignment = Alignment(horizontal="center", vertical="center")
    status_cell.border = thin_border
    
    # Categorization Table
    ws_dash['A7'] = "Category Summary Table"
    ws_dash['A7'].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    
    cat_headers = ["Category", "Transaction Count", "Debits (Out)", "Credits (In)", "Net Impact"]
    for idx, h in enumerate(cat_headers):
        cell = ws_dash.cell(row=8, column=idx+1)
        cell.value = h
        cell.font = Font(name="Calibri", size=11, bold=True, color=white)
        cell.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
        cell.alignment = Alignment(horizontal="left" if idx == 0 else "right", vertical="center")
        cell.border = thin_border

    # Perform categorization logic
    cat_data = df_reconciled.copy()
    cat_data['category'] = cat_data['description'].apply(categorize_transaction)
    
    cat_summary = cat_data.groupby('category').agg(
        count=('description', 'count'),
        debit_sum=('debit', 'sum'),
        credit_sum=('credit', 'sum')
    ).reset_index()
    
    # Net flow by category (credits - debits)
    cat_summary['net'] = cat_summary['credit_sum'] - cat_summary['debit_sum']
    cat_summary = cat_summary.sort_values(by='debit_sum', ascending=False)
    
    row_idx = 9
    for _, r in cat_summary.iterrows():
        # Category Name
        c1 = ws_dash.cell(row=row_idx, column=1, value=r['category'])
        c1.alignment = Alignment(horizontal="left")
        
        # Count
        c2 = ws_dash.cell(row=row_idx, column=2, value=int(r['count']))
        c2.number_format = "#,##0"
        c2.alignment = Alignment(horizontal="right")
        
        # Debits
        c3 = ws_dash.cell(row=row_idx, column=3, value=float(r['debit_sum']))
        c3.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c3.alignment = Alignment(horizontal="right")
        
        # Credits
        c4 = ws_dash.cell(row=row_idx, column=4, value=float(r['credit_sum']))
        c4.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c4.alignment = Alignment(horizontal="right")
        
        # Net
        c5 = ws_dash.cell(row=row_idx, column=5, value=float(r['net']))
        c5.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c5.alignment = Alignment(horizontal="right")
        if r['net'] < 0:
            c5.font = Font(name="Calibri", color="FF0000")
        elif r['net'] > 0:
            c5.font = Font(name="Calibri", color="008000")
            
        # Alternating row fills
        fill_color = grey_light if row_idx % 2 == 1 else white
        for col in range(1, 6):
            cell = ws_dash.cell(row=row_idx, column=col)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = thin_border
            
        row_idx += 1
        
    # Totals Row
    tot_row = row_idx
    ws_dash.cell(row=tot_row, column=1, value="Total").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=2, value=f"=SUM(B9:B{tot_row-1})").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=2).number_format = "#,##0"
    ws_dash.cell(row=tot_row, column=2).alignment = Alignment(horizontal="right")
    
    ws_dash.cell(row=tot_row, column=3, value=f"=SUM(C9:C{tot_row-1})").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=3).number_format = "$#,##0.00;($#,##0.00);\"-\""
    ws_dash.cell(row=tot_row, column=3).alignment = Alignment(horizontal="right")
    
    ws_dash.cell(row=tot_row, column=4, value=f"=SUM(D9:D{tot_row-1})").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=4).number_format = "$#,##0.00;($#,##0.00);\"-\""
    ws_dash.cell(row=tot_row, column=4).alignment = Alignment(horizontal="right")
    
    ws_dash.cell(row=tot_row, column=5, value=f"=D{tot_row}-C{tot_row}").font = Font(bold=True)
    ws_dash.cell(row=tot_row, column=5).number_format = "$#,##0.00;($#,##0.00);\"-\""
    ws_dash.cell(row=tot_row, column=5).alignment = Alignment(horizontal="right")
    
    double_bottom_border = Border(
        top=Side(style='thin', color='A0A0A0'),
        bottom=Side(style='double', color='1F4E78')
    )
    for col in range(1, 6):
        cell = ws_dash.cell(row=tot_row, column=col)
        cell.border = double_bottom_border

    # Adjust widths of Dashboard Columns
    for col in ws_dash.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column > 6:
            continue
        max_len = 0
        for cell in col:
            # Skip A1 title length
            if cell.row == 1 or cell.row == 2:
                continue
            val_str = str(cell.value or '')
            if cell.number_format and ('$' in cell.number_format) and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------------------------------------------
    # Sheet 2: Transactions Detail
    # ----------------------------------------------------
    ws_trans = wb.create_sheet(title="Transactions Detail")
    ws_trans.views.sheetView[0].showGridLines = True
    
    headers = ["Date", "Description", "Category", "Withdrawals (Debits)", "Deposits (Credits)", "Statement Balance", "Expected Balance", "Discrepancy", "Status"]
    
    # Headers Styling
    for col_idx, h in enumerate(headers):
        cell = ws_trans.cell(row=1, column=col_idx+1, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color=white)
        cell.fill = PatternFill(start_color=navy_dark, end_color=navy_dark, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_idx in [0, 8] else "left" if col_idx in [1, 2] else "right", vertical="center")
        cell.border = thin_border
        
    for r_idx, row in df_reconciled.iterrows():
        row_num = r_idx + 2
        
        # Formatted cells
        c_date = ws_trans.cell(row=row_num, column=1, value=row['date'])
        c_date.alignment = Alignment(horizontal="center")
        
        c_desc = ws_trans.cell(row=row_num, column=2, value=row['description'])
        
        c_cat = ws_trans.cell(row=row_num, column=3, value=categorize_transaction(row['description']))
        
        c_debit = ws_trans.cell(row=row_num, column=4, value=float(row['debit']))
        c_debit.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c_debit.alignment = Alignment(horizontal="right")
        
        c_credit = ws_trans.cell(row=row_num, column=5, value=float(row['credit']))
        c_credit.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c_credit.alignment = Alignment(horizontal="right")
        
        c_bal = ws_trans.cell(row=row_num, column=6, value=float(row['balance']))
        c_bal.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c_bal.alignment = Alignment(horizontal="right")
        
        c_exp = ws_trans.cell(row=row_num, column=7, value=float(row['expected_balance']))
        c_exp.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c_exp.alignment = Alignment(horizontal="right")
        
        c_disc = ws_trans.cell(row=row_num, column=8, value=float(row['discrepancy']))
        c_disc.number_format = "$#,##0.00;($#,##0.00);\"-\""
        c_disc.alignment = Alignment(horizontal="right")
        if abs(row['discrepancy']) >= 0.05:
            c_disc.font = Font(name="Calibri", bold=True, color="9C0006")
            c_disc.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
        c_status = ws_trans.cell(row=row_num, column=9)
        c_status.alignment = Alignment(horizontal="center")
        if row['reconciled']:
            c_status.value = "OK"
            c_status.font = Font(name="Calibri", bold=True, color="006100")
            c_status.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            c_status.value = "MISMATCH"
            c_status.font = Font(name="Calibri", bold=True, color="9C0006")
            c_status.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
        # Set borders & row background (striping, except where highlighted)
        row_fill = PatternFill(start_color=grey_light, end_color=grey_light, fill_type="solid") if row_num % 2 == 1 else PatternFill(fill_type=None)
        
        for col_idx in range(1, 10):
            cell = ws_trans.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if not cell.fill.fill_type: # Only apply striping if it doesn't already have warning highlights
                if row_num % 2 == 1:
                    cell.fill = row_fill
                    
    # Auto-adjust column widths for transactions detail
    for col in ws_trans.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format and ('$' in cell.number_format) and isinstance(cell.value, (int, float)):
                val_str = f"${cell.value:,.2f}"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_trans.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    wb.save(output_path)
    return output_path
